from app.db.connection import get_conn
from app.schemas.routes import Route, RouteFilter, RouteExportRequest, Organization
from typing import List, Optional, Dict, Any
from fastapi import HTTPException, status
import logging
from datetime import datetime
import csv
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

def get_organizations() -> List[Organization]:
    """Fetch all available organizations for dropdown selection"""
    try:
        with next(get_conn()) as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT id, organization_name, title FROM Organizations ORDER BY organization_name")
                rows = cur.fetchall()
                return [Organization(id=row[0], organization_name=row[1], title=row[2]) for row in rows]
    except Exception as e:
        logger.error(f"Error fetching organizations: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to fetch organizations")

def get_routes_with_filters(route_filter: RouteFilter) -> Dict[str, Any]:
    """Fetch routes with filtering and pagination using JOINs"""
    try:
        with next(get_conn()) as conn:
            with conn.cursor() as cur:
                # Build dynamic query with JOINs to resolve names
                base_query = """
                    SELECT 
                        ROW_NUMBER() OVER (ORDER BY r.tracking_time DESC) as sn,
                        t.imei as terminal_id,
                        to.name as tracking_object_name,
                        r.tracking_time,
                        r.gps_x,
                        r.gps_y
                    FROM Route_List r
                    LEFT JOIN Tags t ON r.terminal_id = t.imei
                    LEFT JOIN Tracking_Objects to ON t.imei = to.mac
                    LEFT JOIN Organizations o ON t.organization = o.organization_name
                    WHERE o.organization_name = %s
                """
                count_query = """
                    SELECT COUNT(*)
                    FROM Route_List r
                    LEFT JOIN Tags t ON r.terminal_id = t.imei
                    LEFT JOIN Organizations o ON t.organization = o.organization_name
                    WHERE o.organization_name = %s
                """
                params = [route_filter.organization]
                
                if route_filter.start_time:
                    base_query += " AND r.tracking_time >= %s"
                    count_query += " AND r.tracking_time >= %s"
                    params.append(route_filter.start_time)
                
                if route_filter.end_time:
                    base_query += " AND r.tracking_time <= %s"
                    count_query += " AND r.tracking_time <= %s"
                    params.append(route_filter.end_time)
                
                # Get total count
                cur.execute(count_query, params)
                total_count = cur.fetchone()[0]
                
                # Add pagination
                base_query += " ORDER BY r.tracking_time DESC LIMIT %s OFFSET %s"
                offset = (route_filter.page - 1) * route_filter.page_size
                params.extend([route_filter.page_size, offset])
                
                cur.execute(base_query, params)
                rows = cur.fetchall()
                
                routes = [
                    Route(
                        sn=row[0],
                        terminal_id=row[1] or "Unknown",
                        tracking_object=row[2] or "Unknown",
                        tracking_time=row[3],
                        gps_x=row[4],
                        gps_y=row[5]
                    ) for row in rows
                ]
                
                return {
                    "routes": routes,
                    "pagination": {
                        "page": route_filter.page,
                        "page_size": route_filter.page_size,
                        "total_records": total_count
                    },
                    "message": "No data available" if total_count == 0 else None
                }
    except Exception as e:
        logger.error(f"Error fetching routes with filters: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to fetch routes")

def export_routes_to_csv(route_filter: RouteFilter) -> str:
    """Export routes to CSV format"""
    try:
        with next(get_conn()) as conn:
            with conn.cursor() as cur:
                # Build query similar to get_routes_with_filters but without pagination
                base_query = """
                    SELECT 
                        ROW_NUMBER() OVER (ORDER BY r.tracking_time DESC) as sn,
                        t.imei as terminal_id,
                        to.name as tracking_object_name,
                        r.tracking_time,
                        r.gps_x,
                        r.gps_y
                    FROM Route_List r
                    LEFT JOIN Tags t ON r.terminal_id = t.imei
                    LEFT JOIN Tracking_Objects to ON t.imei = to.mac
                    LEFT JOIN Organizations o ON t.organization = o.organization_name
                    WHERE o.organization_name = %s
                """
                params = [route_filter.organization]
                
                if route_filter.start_time:
                    base_query += " AND r.tracking_time >= %s"
                    params.append(route_filter.start_time)
                
                if route_filter.end_time:
                    base_query += " AND r.tracking_time <= %s"
                    params.append(route_filter.end_time)
                
                base_query += " ORDER BY r.tracking_time DESC"
                
                cur.execute(base_query, params)
                rows = cur.fetchall()
                
                # Create CSV content
                output = io.StringIO()
                writer = csv.writer(output)
                
                # Write header
                writer.writerow([
                    'SN', 'Terminal ID', 'Tracking Object', 'Tracking Time', 'GPS X', 'GPS Y'
                ])
                
                # Write data
                for row in rows:
                    writer.writerow([
                        row[0], row[1] or "Unknown", row[2] or "Unknown", 
                        row[3], row[4], row[5]
                    ])
                
                csv_content = output.getvalue()
                output.close()
                
                logger.info(f"Exported {len(rows)} routes to CSV for organization: {route_filter.organization}")
                return csv_content
                
    except Exception as e:
        logger.error(f"Error exporting routes to CSV: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to export routes")

def export_routes_to_xlsx(route_filter: RouteFilter) -> bytes:
    """Export routes to XLSX format using openpyxl"""
    try:
        with next(get_conn()) as conn:
            with conn.cursor() as cur:
                base_query = """
                    SELECT 
                        ROW_NUMBER() OVER (ORDER BY r.tracking_time DESC) as sn,
                        t.imei as terminal_id,
                        to.name as tracking_object_name,
                        r.tracking_time,
                        r.gps_x,
                        r.gps_y
                    FROM Route_List r
                    LEFT JOIN Tags t ON r.terminal_id = t.imei
                    LEFT JOIN Tracking_Objects to ON t.imei = to.mac
                    LEFT JOIN Organizations o ON t.organization = o.organization_name
                    WHERE o.organization_name = %s
                """
                params = [route_filter.organization]
                
                if route_filter.start_time:
                    base_query += " AND r.tracking_time >= %s"
                    params.append(route_filter.start_time)
                
                if route_filter.end_time:
                    base_query += " AND r.tracking_time <= %s"
                    params.append(route_filter.end_time)
                
                base_query += " ORDER BY r.tracking_time DESC"
                
                cur.execute(base_query, params)
                rows = cur.fetchall()
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Routes"
                
                # Write header
                headers = ['SN', 'Terminal ID', 'Tracking Object', 'Tracking Time', 'GPS X', 'GPS Y']
                ws.append(headers)
                
                # Write data
                for row in rows:
                    ws.append([
                        row[0], row[1] or "Unknown", row[2] or "Unknown", 
                        row[3], row[4], row[5]
                    ])
                
                # Auto-size columns
                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except Exception:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2
                
                output = io.BytesIO()
                wb.save(output)
                xlsx_content = output.getvalue()
                output.close()
                
                logger.info(f"Exported {len(rows)} routes to XLSX for organization: {route_filter.organization}")
                return xlsx_content
    except Exception as e:
        logger.error(f"Error exporting routes to XLSX: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to export routes")

def get_route_by_id(route_id: int) -> Optional[Route]:
    """Fetch a specific route by ID"""
    try:
        with next(get_conn()) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT 
                        ROW_NUMBER() OVER (ORDER BY r.tracking_time DESC) as sn,
                        t.imei as terminal_id,
                        to.name as tracking_object_name,
                        r.tracking_time,
                        r.gps_x,
                        r.gps_y
                    FROM Route_List r
                    LEFT JOIN Tags t ON r.terminal_id = t.imei
                    LEFT JOIN Tracking_Objects to ON t.imei = to.mac
                    WHERE r.sn = %s
                    """,
                    (route_id,)
                )
                row = cur.fetchone()
                if row:
                    return Route(
                        sn=row[0],
                        terminal_id=row[1] or "Unknown",
                        tracking_object=row[2] or "Unknown",
                        tracking_time=row[3],
                        gps_x=row[4],
                        gps_y=row[5]
                    )
                return None
    except Exception as e:
        logger.error(f"Error fetching route {route_id}: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to fetch route")

def get_route_statistics(organization: str, start_time: datetime, end_time: datetime) -> Dict[str, Any]:
    """Get route statistics for analytics (future-ready)"""
    try:
        with next(get_conn()) as conn:
            with conn.cursor() as cur:
                # Basic statistics query
                stats_query = """
                    SELECT 
                        COUNT(*) as total_routes,
                        COUNT(DISTINCT r.terminal_id) as unique_terminals,
                        MIN(r.tracking_time) as earliest_time,
                        MAX(r.tracking_time) as latest_time
                    FROM Route_List r
                    LEFT JOIN Tags t ON r.terminal_id = t.imei
                    LEFT JOIN Organizations o ON t.organization = o.organization_name
                    WHERE o.organization_name = %s
                    AND r.tracking_time >= %s
                    AND r.tracking_time <= %s
                """
                
                cur.execute(stats_query, (organization, start_time, end_time))
                row = cur.fetchone()
                
                if row:
                    return {
                        "total_routes": row[0],
                        "unique_terminals": row[1],
                        "earliest_time": row[2],
                        "latest_time": row[3],
                        "time_range_days": (end_time - start_time).days if row[2] and row[3] else 0
                    }
                return {
                    "total_routes": 0,
                    "unique_terminals": 0,
                    "earliest_time": None,
                    "latest_time": None,
                    "time_range_days": 0
                }
                
    except Exception as e:
        logger.error(f"Error fetching route statistics: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to fetch route statistics") 